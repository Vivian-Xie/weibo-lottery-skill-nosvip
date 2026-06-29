#!/usr/bin/env python
# encoding: utf-8
"""
微博抽奖：读取 crawl.py 产出的 jsonl -> 话题过滤(可选) -> 按用户ID去重 -> 随机抽 N 人 -> 输出名单。

特性:
- 按 user._id 去重（同一人多次转发/评论只算一次参与）。
- 可选话题过滤 --tags：仅保留文案 content 同时包含所有指定话题的记录。
- 可选排除名单 --exclude-uid：从参与池里剔除指定用户ID（常用于去掉博主本人/小号/工作人员），可多个。
- 加密级随机 random.SystemRandom，公平不可预测。
- 输出格式 --format csv|xlsx（xlsx 需 openpyxl）。
- --winners-out 指定中奖名单文件；提供 --pool-out 则额外导出去重后的完整参与名单。

用法:
    python lottery.py --jsonl ./output/repost_xxx.jsonl --n 1 \
        --label "转发抽一位·大月卡" \
        --tags "#刃恒99#" "#千冶刃值得#" \
        --exclude-uid 5634207347 \
        --format xlsx \
        --winners-out ./output/中奖名单.xlsx \
        --pool-out ./output/有效参与名单.xlsx
"""
import argparse
import csv
import datetime
import json
import os
import random
import re


# 匹配一对 # 包裹的完整微博话题，话题名内部不含 # 和空白
_TOPIC_RE = re.compile(r"#([^#\s]+)#")


def normalize_tag(t):
    """把用户传入的 tag 规范化为话题名（去掉首尾 # 和空白）。
    #cp# -> cp ; #cp -> cp ; cp# -> cp ; cp -> cp
    """
    return t.strip().strip("#").strip()


def extract_topics(content):
    """从文案中解析出所有【完整话题】token 的集合（话题名，不含 #）。
    例如 '#cp# 甜 #cp99#' -> {'cp', 'cp99'}
    """
    return {m.group(1) for m in _TOPIC_RE.finditer(content or "")}


def load_pool(jsonl_path, tags, exclude_uids=None):
    # 规范化要求的话题名，做完整精确匹配（不模糊、不子串）
    required = {normalize_tag(t) for t in tags if normalize_tag(t)}
    # 需要从参与池剔除的用户ID（如博主本人/小号），统一转字符串精确比对
    excluded = {str(x).strip() for x in (exclude_uids or []) if str(x).strip()}
    total, valid_cnt, excluded_cnt, seen = 0, 0, 0, {}
    for line in open(jsonl_path, "rt", encoding="utf-8"):
        line = line.strip()
        if not line:
            continue
        total += 1
        try:
            rec = json.loads(line)
        except Exception:  # noqa
            continue
        content = rec.get("content", "") or ""
        if required:
            topics = extract_topics(content)
            # 必须把每个要求的话题作为【完整话题】出现，#cp# 绝不命中 #cp99#
            if not required.issubset(topics):
                continue
        valid_cnt += 1
        u = rec.get("user", {}) or {}
        uid = str(u.get("_id", ""))
        if not uid:
            continue
        if uid in excluded:
            excluded_cnt += 1
            continue
        if uid not in seen:
            seen[uid] = {
                "uid": uid,
                "nick_name": u.get("nick_name", ""),
                "ip_location": rec.get("ip_location", "") or "",
                "created_at": rec.get("created_at", ""),
                "home": f"https://weibo.com/u/{uid}",
                "content": content,
            }
    return total, valid_cnt, excluded_cnt, list(seen.values())


def write_csv(path, rows, header):
    with open(path, "wt", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rows:
            w.writerow(r)


def write_winners_xlsx(path, label, winners, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "中奖名单"
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "微博抽奖 · 中奖名单"
    t.font = Font(bold=True, size=14, color="C00000")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.append(["奖项", "昵称", "用户ID", "IP属地", "主页链接"])
    hf = PatternFill("solid", fgColor="C00000")
    for c in ws[2]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center")
    for w in winners:
        ws.append([label, w["nick_name"], w["uid"], w["ip_location"], w["home"]])
    start = 3 + len(winners)
    for j, line in enumerate(meta):
        ws.cell(row=start + j, column=1, value=line)
    for col, wd in zip("ABCDE", [22, 26, 16, 10, 32]):
        ws.column_dimensions[col].width = wd
    wb.save(path)


def write_pool_xlsx(path, pool):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "有效参与名单"
    ws.append(["序号", "昵称", "用户ID", "IP属地", "首次参与时间", "主页链接"])
    hf = PatternFill("solid", fgColor="4472C4")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hf
        c.alignment = Alignment(horizontal="center")
    for i, p in enumerate(pool, 1):
        ws.append([i, p["nick_name"], p["uid"], p["ip_location"], p["created_at"], p["home"]])
    for col, wd in zip("ABCDEF", [6, 26, 16, 10, 20, 32]):
        ws.column_dimensions[col].width = wd
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--jsonl", required=True)
    ap.add_argument("--n", type=int, default=1, help="抽取人数")
    ap.add_argument("--label", default="抽一位", help="奖项名称")
    ap.add_argument("--tags", nargs="*", default=[], help="必须同时包含的话题（可多个）")
    ap.add_argument("--exclude-uid", nargs="*", default=[],
                    help="从参与池剔除的用户ID（如博主本人/小号/工作人员，可多个）")
    ap.add_argument("--format", choices=["csv", "xlsx"], default="csv")
    ap.add_argument("--winners-out", required=True)
    ap.add_argument("--pool-out", default="")
    args = ap.parse_args()

    total, valid_cnt, excluded_cnt, pool = load_pool(args.jsonl, args.tags, args.exclude_uid)
    excl_note = f" | 剔除指定UID {excluded_cnt} 条" if args.exclude_uid else ""
    print(f"[统计] 原始 {total} 条 | 满足条件 {valid_cnt} 条{excl_note} | 去重后参与 {len(pool)} 人")
    if not pool:
        print("[错误] 参与池为空，无法抽奖")
        raise SystemExit(2)
    n = min(args.n, len(pool))
    rng = random.SystemRandom()
    winners = rng.sample(pool, n)

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta = [
        f"开奖时间：{now}",
        f"有效参与条件：{'文案同时含 ' + ' '.join(args.tags) if args.tags else '无（全部转发/评论者）'}",
        f"原始记录数：{total}",
        f"满足条件记录数：{valid_cnt}",
        f"排除名单（UID {('、'.join(map(str, args.exclude_uid))) if args.exclude_uid else '无'}）：剔除 {excluded_cnt} 条",
        f"去重后参与人数：{len(pool)}",
        "抽取方式：random.SystemRandom 加密级随机，公平不可预测",
    ]

    if args.format == "xlsx":
        write_winners_xlsx(args.winners_out, args.label, winners, meta)
        if args.pool_out:
            write_pool_xlsx(args.pool_out, pool)
    else:
        rows = [[args.label, w["nick_name"], w["uid"], w["ip_location"], w["home"]] for w in winners]
        write_csv(args.winners_out, rows, ["奖项", "昵称", "用户ID", "IP属地", "主页链接"])
        if args.pool_out:
            prows = [[i, p["nick_name"], p["uid"], p["ip_location"], p["created_at"], p["home"]]
                     for i, p in enumerate(pool, 1)]
            write_csv(args.pool_out, prows, ["序号", "昵称", "用户ID", "IP属地", "首次参与时间", "主页链接"])

    print("\n========= 开奖结果 =========")
    for w in winners:
        print(f"【{args.label}】{w['nick_name']}  (UID {w['uid']}, {w['ip_location']})  {w['home']}")
    print("============================")
    print(f"[产出] 中奖名单: {os.path.abspath(args.winners_out)}")
    if args.pool_out:
        print(f"[产出] 参与名单: {os.path.abspath(args.pool_out)}")


if __name__ == "__main__":
    main()
